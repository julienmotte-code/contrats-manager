"""
Import et lecture des recapitulatifs de marge brute (Excel historiques).
Source : onglet 'recapitulatif' de chaque "CALCUL MARGE BRUTE <annee>.xlsx".
Reproduit le tableau tel quel : familles (lignes) x 12 mois + total. Pas de recalcul metier.
"""
import os, re, glob
from decimal import Decimal
from openpyxl import load_workbook
from app.models.models import CaRecapExcel

MOIS_LABELS = ["Janvier","Fevrier","Mars","Avril","Mai","Juin",
               "Juillet","Aout","Septembre","Octobre","Novembre","Decembre"]

# Accepte espaces OU underscores, ignore les fichiers temporaires Excel (~$...)
GLOB_PATTERNS = ["CALCUL*MARGE*BRUTE*.xlsx"]

def _to_dec(v):
    if v is None or v == "": return None
    try: return Decimal(str(round(float(v), 2)))
    except Exception: return None

def _is_code(v):
    return isinstance(v,(int,float)) and 70000000 <= int(v) <= 70999999

def _annee_from_name(path):
    m = re.search(r"(20\d{2})", os.path.basename(path))
    return int(m.group(1)) if m else None

def _sheet_recap(wb):
    for name in wb.sheetnames:
        if name.strip().lower() == "recapitulatif":
            return wb[name]
    return None

def _lister_fichiers(dossier):
    out = []
    for pat in GLOB_PATTERNS:
        out += glob.glob(os.path.join(dossier, pat))
    # exclure les fichiers de verrou Excel "~$..."
    out = [p for p in out if not os.path.basename(p).startswith("~$")]
    return sorted(set(out))

def _parse_recap(path):
    annee = _annee_from_name(path)
    wb = load_workbook(path, data_only=True)
    ws = _sheet_recap(wb)
    if ws is None:
        wb.close(); return annee, []
    rows = list(ws.iter_rows(min_row=1, max_row=40, max_col=16, values_only=True))
    col_jan = None; hdr_idx = None
    for i, row in enumerate(rows):
        for j, v in enumerate(row):
            if isinstance(v, str) and v.strip().upper().startswith("JANVIER"):
                hdr_idx, col_jan = i, j; break
        if col_jan is not None: break
    if col_jan is None:
        wb.close(); return annee, []
    familles = []
    for row in rows[hdr_idx+1:]:
        a = row[0] if len(row) > 0 else None
        b = row[1] if len(row) > 1 else None
        if isinstance(a, str) and a.strip().upper().startswith("TOTAL"):
            break
        if _is_code(a):
            mois = [_to_dec(row[col_jan+k]) if col_jan+k < len(row) else None for k in range(12)]
            total = sum((m for m in mois if m is not None), Decimal("0"))
            familles.append({"code": str(int(a)), "libelle": (str(b).strip() if b else ""),
                             "mois": mois, "total": Decimal(str(round(float(total),2)))})
    wb.close()
    return annee, familles

def importer_recap_excel(db, dossier):
    fichiers = _lister_fichiers(dossier)
    resume = []
    for path in fichiers:
        annee, familles = _parse_recap(path)
        if not annee or not familles:
            resume.append({"fichier": os.path.basename(path), "annee": annee, "familles": 0,
                           "total": 0.0, "statut": "VIDE/ignore"}); continue
        db.query(CaRecapExcel).filter(CaRecapExcel.annee == annee).delete()
        for ordre, f in enumerate(familles):
            db.add(CaRecapExcel(annee=annee, ordre=ordre, code_compte=f["code"],
                   famille_libelle=f["libelle"], total_ht=f["total"],
                   **{f"m{m:02d}": f["mois"][m-1] for m in range(1,13)}))
        db.commit()
        resume.append({"fichier": os.path.basename(path), "annee": annee, "familles": len(familles),
                       "total": float(sum((f["total"] for f in familles), Decimal("0"))), "statut": "OK"})
    return resume

def annees_disponibles(db):
    return [a for (a,) in db.query(CaRecapExcel.annee).distinct()
            .order_by(CaRecapExcel.annee.desc()).all()]

def get_recap(db, annee):
    lignes = db.query(CaRecapExcel).filter(CaRecapExcel.annee == annee).order_by(CaRecapExcel.ordre).all()
    familles = []; totaux_mois = [Decimal("0")]*12
    for L in lignes:
        mois = [getattr(L, f"m{m:02d}") for m in range(1,13)]
        for k in range(12):
            if mois[k] is not None: totaux_mois[k] += mois[k]
        familles.append({"code": L.code_compte, "famille": L.famille_libelle,
                         "mois": [float(x) if x is not None else None for x in mois],
                         "total": float(L.total_ht or 0)})
    total_annee = float(sum((L.total_ht or Decimal("0") for L in lignes), Decimal("0")))
    return {"annee": annee, "mois_labels": MOIS_LABELS, "familles": familles,
            "totaux_mois": [float(x) for x in totaux_mois], "total_annee": round(total_annee,2)}
