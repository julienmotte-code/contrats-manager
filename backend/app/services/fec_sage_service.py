"""
Conversion d'un export FEC Karlia (.xlsx) vers le format d'import Sage 100
(blocs #MECG). Stateless : aucune dépendance DB.

Correspondances validées avec la compta SGI (plan Sage 100cloud) sur deux
exports réels (avril et mai 2026).
"""
from __future__ import annotations

import io
import logging
from collections import Counter, OrderedDict
from dataclasses import dataclass
from typing import Dict, List, Tuple

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Correspondances comptes Karlia -> Sage (le reste passe à l'identique)
COMPTE_MAP: Dict[str, str] = {
    "401": "40100000",      # Fournisseurs
    "411": "41100000",      # Clients
    "471": "47100000",      # Compte d'attente
    "512003": "51211000",   # C/C Eurocompte -> Crédit Mutuel du Nord
    "445661": "44566200",   # TVA déductible 20%
    "445711": "44571200",   # TVA collectée 20%
}
# Journaux Karlia -> Sage
JOURNAL_MAP: Dict[str, str] = {"ACH": "ACH", "VTE": "VTE", "BQ": "CMN"}
JOURNAL_BANQUE_KARLIA = "BQ"
COLLECTIFS = ("401", "411")
COLONNES_REQUISES = ["JournalCode", "EcritureDate", "CompteNum", "PieceDate",
                     "EcritureLib", "Debit", "Credit"]
_TAIL = ["", "", "0", "__SENS__", "__MONTANT__", "", "", "", "0", "0", "0",
         "", "", "0", "0", "0", "", "", "", "", "0", "", "", "", "0"]
ENTETE_SAGE = ["#FLG 001", "#VER 15"]
ENCODAGE_SAGE = "latin-1"
LIBELLE_MAX = 35


class FecIntegriteError(ValueError):
    """Erreur de contrôle d'intégrité du fichier FEC d'entrée."""


@dataclass
class RecapJournal:
    code_karlia: str
    code_sage: str
    nb_ecritures: int
    total_debit: float
    total_credit: float


@dataclass
class RecapConversion:
    nb_lignes: int
    total_debit: float
    total_credit: float
    equilibre: bool
    periode_min: str
    periode_max: str
    journaux: List[RecapJournal]
    comptes_utilises: List[str]
    banque_incluse: bool


def _norm8(v) -> str:
    if v in (None, ""):
        return ""
    s = str(int(v)) if isinstance(v, (int, float)) else str(v).strip()
    return s.zfill(8)


def _to_ddmmyy(v) -> str:
    s = _norm8(v)
    return s[0:2] + s[2:4] + s[6:8] if s else ""


def _map_compte(cn: str) -> str:
    cn = str(cn).strip().split(".")[0]
    return COMPTE_MAP.get(cn, cn)


def _map_aux(auxnum: str, comptenum: str) -> str:
    a = str(auxnum).strip()
    if not a:
        return ""
    for col in COLLECTIFS:
        if a.startswith(col):
            return a[len(col):]
    return a


def _fmt(x: float) -> str:
    return f"{abs(float(x)):.2f}"


def _sans_tiret_initial(s: str) -> str:
    s = str(s).strip()
    return s[1:].lstrip() if s.startswith("-") else s


def convertir_fec_vers_sage(contenu_xlsx: bytes,
                            inclure_banque: bool = False) -> Tuple[bytes, RecapConversion]:
    """Convertit un FEC Karlia (.xlsx) en fichier d'import Sage (latin-1, CRLF)."""
    try:
        wb = load_workbook(io.BytesIO(contenu_xlsx), data_only=True, read_only=True)
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as exc:
        raise FecIntegriteError(
            f"Fichier illisible : ce n'est pas un .xlsx valide ({exc})."
        )

    if not rows:
        raise FecIntegriteError("Fichier vide : aucune ligne détectée.")

    entetes = [("" if c is None else str(c).strip()) for c in rows[0]]
    manquantes = [c for c in COLONNES_REQUISES if c not in entetes]
    if manquantes:
        raise FecIntegriteError(
            "En-têtes FEC manquants : " + ", ".join(manquantes)
            + ". Le fichier ne semble pas être un export FEC Karlia."
        )

    ix = {h: i for i, h in enumerate(entetes)}
    data = rows[1:]
    if not data:
        raise FecIntegriteError(
            "Fichier sans écriture : seule la ligne d'en-têtes est présente "
            "(export vide ou mauvaise plage de dates)."
        )

    def cell(row, name):
        i = ix.get(name)
        if i is None or i >= len(row):
            return ""
        v = row[i]
        return "" if v is None else v

    blocs: List[List[str]] = []
    comptes_inconnus = set()
    comptes_utilises = OrderedDict()
    jr_count, jr_debit, jr_credit = Counter(), Counter(), Counter()
    total_debit = total_credit = 0.0
    dates = []

    for num_ligne, row in enumerate(data, start=2):
        jc = str(cell(row, "JournalCode")).strip()
        if not jc:
            continue
        if not inclure_banque and jc == JOURNAL_BANQUE_KARLIA:
            continue

        cn = str(cell(row, "CompteNum")).strip()
        if not cn:
            raise FecIntegriteError(
                f"Ligne {num_ligne} : compte général absent (fichier tronqué ?)."
            )
        cmpt = _map_compte(cn)
        if cn not in COMPTE_MAP and not (cmpt.isdigit() and len(cmpt) == 8):
            comptes_inconnus.add(cn)

        d = float(cell(row, "Debit") or 0)
        c = float(cell(row, "Credit") or 0)
        sens = "0" if d > 0 else "1"
        montant = _fmt(d if d > 0 else c)
        aux = _map_aux(cell(row, "CompteAuxNum"), cn)
        de = cell(row, "EcritureDate")
        if de != "":
            dates.append(_norm8(de))

        b = [""] * 38
        b[0] = JOURNAL_MAP.get(jc, jc)
        b[1] = _to_ddmmyy(de)
        b[2] = _to_ddmmyy(cell(row, "PieceDate"))
        b[3] = str(cell(row, "EcritureNum")).strip()
        b[4] = _sans_tiret_initial(cell(row, "PieceRef"))
        b[6] = cmpt
        b[8] = aux
        b[10] = _sans_tiret_initial(cell(row, "EcritureLib"))[:LIBELLE_MAX]
        b[11] = "0"
        b[12] = _to_ddmmyy(cell(row, "DateRglt") or cell(row, "PieceDate"))
        for i, tv in enumerate(_TAIL, start=13):
            b[i] = sens if tv == "__SENS__" else (montant if tv == "__MONTANT__" else tv)
        blocs.append(b)

        comptes_utilises[cmpt] = True
        jr_count[jc] += 1
        jr_debit[jc] += d
        jr_credit[jc] += c
        total_debit += d
        total_credit += c

    if comptes_inconnus:
        raise FecIntegriteError(
            "Comptes non reconnus (ni mappés, ni au format 8 chiffres) : "
            + ", ".join(sorted(comptes_inconnus)) + ". Mapping à compléter."
        )
    if not blocs:
        raise FecIntegriteError(
            "Aucune écriture à exporter (seul le journal de banque est présent ? "
            "cochez « inclure la banque » si besoin)."
        )

    equilibre = abs(total_debit - total_credit) < 0.01
    if not equilibre:
        raise FecIntegriteError(
            f"Écritures déséquilibrées : débit {total_debit:.2f} != crédit "
            f"{total_credit:.2f} (écart {total_debit - total_credit:+.2f}). "
            "Fichier probablement tronqué."
        )

    lignes = list(ENTETE_SAGE)
    for b in blocs:
        lignes.append("#MECG")
        lignes.extend(b)
    octets = ("\r\n".join(lignes) + "\r\n").encode(ENCODAGE_SAGE, errors="replace")

    dates.sort()
    fmtd = lambda s: f"{s[0:2]}/{s[2:4]}/{s[4:8]}" if s else ""
    recap = RecapConversion(
        nb_lignes=len(blocs),
        total_debit=round(total_debit, 2),
        total_credit=round(total_credit, 2),
        equilibre=equilibre,
        periode_min=fmtd(dates[0]) if dates else "",
        periode_max=fmtd(dates[-1]) if dates else "",
        journaux=[RecapJournal(jc, JOURNAL_MAP.get(jc, jc), jr_count[jc],
                               round(jr_debit[jc], 2), round(jr_credit[jc], 2))
                  for jc in sorted(jr_count)],
        comptes_utilises=sorted(comptes_utilises),
        banque_incluse=inclure_banque,
    )
    return octets, recap
